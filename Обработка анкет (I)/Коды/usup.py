import openpyxl

# открыли книжку
book = openpyxl.open(".\mirst.xlsx", read_only=True)

# взяли страничку
sheet = book.worksheets[3]

lst = []

cnt = 0
lst1 = [0, 0, 0, 0, 0, 0]
for col in range (1, sheet.max_column+1):
    if col>=10 and col%7==3:
        for row in range (1, 42):
            if row==5:
                    if cnt>=1:
                        lst.append(lst1)    
                    cnt+=1
                    lst1 = [0, 0, 0, 0, 0, 0]
            if row>=5 and sheet.cell(row, col).value!= None:
                for n in range(1, 6):
                    if sheet.cell(row, col-n).value!= None:
                        ans = sheet.cell(row, col-n).value
                    
                
                if sheet.cell(row, 3).value == "Герои" :
                    lst1[0] = lst1[0]+ans
                    print(1)
                elif sheet.cell(row, 3).value == "Дети":
                    lst1[1] = lst1[1]+ans
                    print(1)
                elif sheet.cell(row, 3).value == "Животные ":
                    lst1[2] = lst1[2]+ans
                    print(1)
                elif sheet.cell(row, 3).value == "незнакомые ":
                    lst1[3] = lst1[3]+ans
                    print(1)
                elif sheet.cell(row, 3).value == "Родители ":
                    lst1[4] = lst1[4]+ans
                    print(1)
                elif sheet.cell(row, 3).value == "Старики":
                    lst1[5] = lst1[5]+ans
                    print(1)
                
                
lst.append(lst1)
print(lst)
print(cnt)
# m[[5, 1, 6, 6, 11, 4], [7, 1, 7, 5, 11, 8], [3, 1, 9, 7, 9, 7], [3, 1, 9, 7, 9, 6], [9, 2, 6, 6, 14, 8], [5, 1, 7, 5, 7, 5], [4, 4, 9, 10, 4, 7], [3, 2, 9, 5, 8, 6], [6, 1, 6, 5, 6, 6], [3, 5, 7, 9, 11, 9], [5, 2, 7, 7, 10, 8], [6, 1, 8, 5, 7, 5], [5, 3, 9, 8, 11, 9], [5, 2, 8, 6, 7, 5], [5, 2, 11, 8, 12, 7],
#d [[5, 2, 14, 9, 11, 7], [7, 4, 6, 9, 13, 9], [6, 1, 4, 4, 10, 4], [4, 3, 7, 8, 6, 6], [7, 3, 6, 10, 10, 9], [4, 5, 3, 8, 9, 7], [6, 4, 10, 8, 4, 9], [5, 2, 6, 5, 8, 6], [8, 2, 5, 11, 12, 10], [3, 2, 9, 7, 9, 8], [5, 2, 5, 10, 11, 8], [11, 4, 13, 12, 13, 7], [8, 3, 11, 10, 9, 7], [5, 1, 13, 10, 6, 10], [8, 4, 7, 13, 10, 6]]