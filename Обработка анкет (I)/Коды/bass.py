import openpyxl

# открыли книжку
book = openpyxl.open(".\dirst.xlsx", read_only=True)

# взяли страничку
sheet = book.worksheets[1]

lst = []

cnt = 0
lst1 = [0, 0, 0]
for col in range (1, sheet.max_column+1):
    if col>=4 and col%2==0:
        for row in range (1, sheet.max_row+1):
            if row==4:
                    if cnt>=1:
                        lst.append(lst1)    
                    cnt+=1
                    lst1 = [0, 0, 0]
            if row>=4 and sheet.cell(row, col).value!= None:
                if sheet.cell(row, 3).value == "НА СЕБЯ" :
                    lst1[0] = lst1[0]+sheet.cell(row, col).value
                elif sheet.cell(row, 3).value == "ДЕЛО":
                    lst1[1] = lst1[1]+sheet.cell(row, col).value
                elif sheet.cell(row, 3).value == "ОБЩЕНИЕ":
                    lst1[2] = lst1[2]+sheet.cell(row, col).value
lst.append(lst1)
print(lst1)
print(cnt)
