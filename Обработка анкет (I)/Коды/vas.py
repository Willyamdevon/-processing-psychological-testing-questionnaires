import openpyxl
import pandas as pd

# открыли книжку
book = openpyxl.open(".\mirst.xlsx", read_only=True)

# взяли страничку
sheet = book.worksheets[4 ]

lst = []
cnt = 0
lst1 = [0, 0, 0, 0, 0, 0, 0, 0, 0] 
for col in range (1, sheet.max_column+1):
    
    if col>=5 and col%2==1:
        for row in range (1, sheet.max_row+1):
            
            if row==3:
                    if cnt>=1:
                        lst.append(lst1)    
                    cnt+=1
                    lst1 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
            if row>=3 and sheet.cell(row, col).value!= None and sheet.cell(row, 2).value!= None:   
                
                if sheet.cell(row, 2).value == "Возможность заботиться о здоровье своих близких": lst1[0] = lst1[0]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Возможность заботиться о своем здоровье": lst1[1] = lst1[1]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Возможность оказывать влияние на других людей": lst1[2] = lst1[2]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Доступность медикаментов": lst1[3] = lst1[3]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Желание лечить людей": lst1[4] = lst1[4]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Желание облегчить страдания тяжелобольных, стариков и детей": lst1[5] = lst1[5]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Престиж профессии и семейные традиции": lst1[6] = lst1[6]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Желание решать научные медицинские проблемы": lst1[7] = lst1[7]+ sheet.cell(row, col).value
                elif sheet.cell(row, 2).value == "Материальная заинтересованность ": lst1[8] = lst1[8]+ sheet.cell(row, col).value

print(lst)
print(cnt)
