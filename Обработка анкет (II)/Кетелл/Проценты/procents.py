import xlwt
import openpyxl

res = xlwt.Workbook(encoding="utf-8")
res_d= res.add_sheet("d",cell_overwrite_ok=True) 
res_m = res.add_sheet("m",cell_overwrite_ok=True)

main = openpyxl.open("numbers.xlsx", read_only=True)
sheet_d = main.worksheets[2]
sheet_m = main.worksheets[3]

for row in range(1, sheet_d.max_row+1):
    if row>=2:
        h = 0
        m = 0
        l = 0
        for col in range(1, sheet_d.max_column+1):
            if col>=2:
                if sheet_d.cell(row, col).value =="LOW": l+=1
                elif sheet_d.cell(row, col).value =="MID": m+=1
                elif sheet_d.cell(row, col).value =="HIG": h+=1
            
        print(l, m, h)
        res_l = round(l*100/15)
        res_m1 = round(m*100/15)
        res_h = round(h*100/15)
        res_d.write(row, 2, res_l)
        res_d.write(row, 3, res_m1)
        res_d.write(row, 4, res_h)
for row in range(1, sheet_m.max_row+1):
    if row>=2:
        h = 0
        m = 0
        l = 0
        for col in range(1, sheet_m.max_column+1):
            if col>=2:
                if sheet_m.cell(row, col).value =="LOW": l+=1
                elif sheet_m.cell(row, col).value =="MID": m+=1
                elif sheet_m.cell(row, col).value =="HIG": h+=1
            
        print(l, m, h)
        res_l = round(l*100/15)
        res_m1 = round(m*100/15)
        res_h = round(h*100/15)
        res_m.write(row, 2, res_l)
        res_m.write(row, 3, res_m1)
        res_m.write(row, 4, res_h)          

res.save("proc.xls")