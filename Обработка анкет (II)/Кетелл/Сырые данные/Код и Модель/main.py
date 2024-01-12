import xlwt
import openpyxl

res = xlwt.Workbook(encoding="utf-8")
res_d= res.add_sheet("1",cell_overwrite_ok=True) 
res_w = res.add_sheet("2",cell_overwrite_ok=True)

maind = openpyxl.open("d.xlsx", read_only=True)
sheet_d = maind.worksheets[5]

mainw = openpyxl.open("wvet.xlsx", read_only=True)
sheet_w = mainw.worksheets[2]

model = openpyxl.open("model.xlsx", read_only=True)
mod = model.worksheets[0]
print("Start")

for col in range(1, 19 ):
    if col>=3:
        for row in range (1, sheet_w.max_row+1):
            if row>=2:
                print(col)
                for col_m in range(2, 12):
                    
                    if type(mod.cell(row, col_m).value) == str:
                        lst = mod.cell(row, col_m).value.split(";")
                        
                        lst1 = []
                        for i in range (len(lst)):
                            lst1.append(int(lst[i]))
                        
                        if sheet_w.cell(row, col).value in lst1:
                            res_d.write(row,col-1, mod.cell(1, col_m).value)
                            
                    elif mod.cell(row, col_m).value!=None:
                        if sheet_w.cell(row, col).value == mod.cell(row, col_m).value:
                            res_d.write(row,col-1, mod.cell(1, col_m).value)
                            
                    
res.save("resul_d.xls")







