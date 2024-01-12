import xlwt
import openpyxl

res = xlwt.Workbook(encoding="utf-8")
res_d= res.add_sheet("d",cell_overwrite_ok=True) 
res_m = res.add_sheet("m",cell_overwrite_ok=True)

main = openpyxl.open("numbers.xlsx", read_only=True)
sheet_d = main.worksheets[0]
sheet_m = main.worksheets[1]


for row in range(1, sheet_d.max_row+1):
    if row>=2:
        for col in range(1, sheet_d.max_column+1):
            if col>=2:
                if sheet_d.cell(row, col).value >6:
                    res_d.write(row, col, "HIG")
                elif sheet_d.cell(row, col).value < 5:
                    res_d.write(row, col, "LOW")
                else: res_d.write(row, col, "MID")
for row in range(1, sheet_m.max_row+1):
    if row>=2:
        for col in range(1, sheet_m.max_column+1):
            if col>=2:
                if sheet_m.cell(row, col).value >6:
                    res_m.write(row, col, "HIG")
                elif sheet_m.cell(row, col).value < 5:
                    res_m.write(row, col, "LOW")
                else: res_m.write(row, col, "MID")


res.save("types.xls")