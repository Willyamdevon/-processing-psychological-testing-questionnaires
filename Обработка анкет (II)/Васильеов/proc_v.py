import openpyxl
import xlwt

book = openpyxl.open("v.xlsx", read_only=True)
sheet = book.worksheets[5]

res = xlwt.Workbook(encoding="utf-8")
res1 = res.add_sheet("1")

for row in range(1, sheet.max_row+1):
    if row>1:
        for col in range(1, sheet.max_column +1):
            if col>1:
                if type(sheet.cell(row, col).value)!= str:
                    a = sheet.cell(row, col).value
                    print(a, row, col)
                    ans = round(a*100 /24)

                    res1.write(row, col, ans)

res.save("res.xls")